import openpyxl as xl
import pandas as pd
from pandas import ExcelWriter

wb = xl.load_workbook('example.xlsx') #load excel in to the python
sheet = wb['Sheet1']
cell = sheet['a1']

a = [] #creat a list to store R S B
for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 2)
    for list in cell.value:
        a.append(list)

a = [x for x in a if x != 'B'] #Delete all the B's

inRow = [] #create another list to store R and S without repeat
last = None
for x in a:
        if last == x and (len(inRow) == 0 or inRow[-1] != x):
            inRow.append(last)
        last = x
print(inRow)

df = pd.DataFrame(inRow) #write the result into a new excel
writer = ExcelWriter('example2.xlsx')
df.to_excel(writer,'Sheet1',index=False)
writer.save()

# res=[]
# for i in a:
#     if a.count(i)>1 and i not in res:
#         res.append(i)
# print(res)

# print(a)
#
# for i in range(len(a)):
#     print(a[i])
# counter = 0
# for i in a:
#     if i == 'S':
#         counter = counter +1
#
# for i in range(counter):
#     a.remove('S')
#
# if a[0] == 'R':
#     for i in range(1, counter+1):
#         a[2*i - 1] = 'S'
#
# print(*a)